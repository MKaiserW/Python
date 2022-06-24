
import sys 
sys.path.append('/home/m335964/miniconda3/lib/python3.9/site-packages')
import openpyxl
import os
import string
import random



filestoremove = []
for filename in os.listdir():
    if ".xlsx" in filename:
        print("\n" + filename )
        filestoremove.append(filename)

print()

for file in filestoremove:        
    wb = openpyxl.load_workbook(file)
    
    wl = wb.sheetnames
    ws = wb[wl[0]]
    

    i = 0
    for r in range(1,ws.max_row+1):
        for c in range(1,ws.max_column+1):
            s = str(ws.cell(r,c).value)
            if s == 'None':
                continue
            
            new_string = s.translate(str.maketrans('', '', string.punctuation))
            
            if new_string.isdigit():
                 
                ws.cell(r,c).value = random.randint(0,1000) 

                i+=1

    wb.save("noTechnicaldata_{}".format(file))
    print("{} cells updated".format(i))